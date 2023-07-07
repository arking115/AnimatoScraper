import pandas as pd
import re
import time
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl
import os
from videolinks import video_links

class Scraper:
    def __init__(self):
        self.driver = webdriver.Chrome()

    def process_views(self, elements):
        views = []
        element_texts = [self.driver.execute_script("return arguments[0].outerHTML;", element) for element in elements]
        for element_text in element_texts:
            view_element = re.search(r'(\d+(?:\.\d+)?)[K|M] views', element_text)
            if view_element:
                views.append(float(view_element.group(1)) * 1000 if 'K' in view_element.group() else float(view_element.group(1)) * 1000000)

        return views
    
    def get_channel_name(self):
        name = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(
                (By.XPATH, "//*[@id='text']"))
            )
        name_text = name.text

        return name_text
    
    def organise_video_views(self):
        list_of_views = self.get_views()
        enumerated_views = [[i + 1, int(view)] for i, view in enumerate(reversed(list_of_views))]

        return enumerated_views

            
    def scroll(self):
        previous_position = None
        
        while True:
            time.sleep(1)
            self.driver.find_element(By.TAG_NAME, value="body").send_keys(Keys.PAGE_DOWN)
            
            current_position = self.driver.execute_script("return window.pageYOffset;")
            
            if current_position == previous_position:
                break
            
            previous_position = current_position
    
    def get_views(self):
        try:
            video_button = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located(
                (By.XPATH, "//*[@id='tabsContent']/tp-yt-paper-tab[2]/div/div[1]"))
            )
            video_button.click()
            self.scroll()
            WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "metadata-line")))
            elements = self.driver.find_elements(By.ID, "metadata-line")

            return self.process_views(elements)
        except Exception as e:
            print("Error occurred while fetching views", str(e))

    def create_excel(self, processed_list, channel_name):
        df = pd.DataFrame(processed_list, columns=["Video Number", "View Count"])
        excel_file = 'Animato Analysis.xlsx'

        if os.path.isfile(excel_file):
            # File exists, load existing workbook
            workbook = openpyxl.load_workbook(excel_file)
        else:
            # File does not exist, create new workbook
            workbook = openpyxl.Workbook()

        # Create a new sheet with channel_name
        if channel_name in workbook.sheetnames:
            # If the sheet already exists, remove it before adding the updated data
            workbook.remove(workbook[channel_name])
        sheet = workbook.create_sheet(channel_name)

        # Append new data to the sheet
        for index, row in df.iterrows():
            sheet.append(row.to_list())

        # Save the updated workbook
        workbook.save(excel_file)


    def get_yt_info(self, video_links):
        for link in video_links:
            self.driver.get(link)
            self.create_excel(self.organise_video_views(), self.get_channel_name())


animato_scraper = Scraper()
animato_scraper.get_yt_info(video_links)
# print(len(video_links))